import openpyxl

class Node:
    def __init__(self, id, row_num, level):
        self.id = id
        self.row_num = row_num
        self.level = level
filterList = []

# counts the unique id in set them in Counts
def set_label_counts_on_count_sheet(sheet_name,
                                    label1,
                                    count1,
                                    label2,
                                    count2,
                                    label3,
                                    count3,
                                    label4,
                                    count4):
    sheet_name.cell(row=1, column=1, value=label1)
    sheet_name.cell(row=1, column=2, value=count1)
    sheet_name.cell(row=2, column=1, value=label2)
    sheet_name.cell(row=2, column=2, value=count2)
    sheet_name.cell(row=3, column=1, value=label3)
    sheet_name.cell(row=3, column=2, value=count3)
    sheet_name.cell(row=4, column=1, value=label4)
    sheet_name.cell(row=4, column=2, value=count4)

# collect the required data by traversing the rows of the source file
def get_required_values_from_one_tab(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    row_list = []
    levels_list = []
    id=""

    for row_number in range(2, sheet.max_row + 1):
        value_in_B = sheet.cell(row=row_number, column=2).value
        value_in_label = sheet.cell(row=row_number, column=29).value

        if row_number == 2:
            # getting the first row as a special case
            id = value_in_B
            row_list.append(row_number)
            levels_list.append(value_in_label)
            continue
# As there were merged cells under one ids, all the data (rows) 
# under one id should be saved as one structure 
        if value_in_B:
            filterList.append(Node(id, row_list, levels_list))
            row_list = []
            levels_list = []

            id = value_in_B
            row_list.append(row_number)
            levels_list.append(value_in_label)
        else:
            row_list.append(row_number)
            levels_list.append(value_in_label)
    # saving the last node as it is collected but not put in into the list
    filterList.append(Node(id, row_list, levels_list))

    for currentNode in filterList:
        print(f"Id:{currentNode.id}, Rows: {currentNode.row_num}, Labels: {currentNode.level}")
    
    workbook.close()

# Doing the copy operations as reqeuired
def copy_row(source_sheet, destination_sheet, row_number):
    source_row = source_sheet[row_number]
    destination_sheet.append([cell.value for cell in source_row])

# Copy rows in different destination tabs based on requirements
def copy_rows_to_dest_tab(file_path,
                      source_sheet, 
                      destination_sheet_1,
                      destination_sheet_2,
                      destination_sheet_3,
                      destination_sheet_4,
                      destination_sheet_5,
                      label_1,
                      label_2,
                      label_3,
                      isColumnNameNotExist):
    workbook = openpyxl.load_workbook(file_path)
    s_sheet = workbook[source_sheet]
    d_sheet1 = workbook[destination_sheet_1]
    d_sheet2 = workbook[destination_sheet_2]
    d_sheet3 = workbook[destination_sheet_3]
    d_sheet4 = workbook[destination_sheet_4]

    countL1 = 0
    countL2 = 0
    countL3 = 0
    countL4 = 0

    if isColumnNameNotExist:
        copy_row(s_sheet, d_sheet1, 1)
        copy_row(s_sheet, d_sheet2, 1)
        copy_row(s_sheet, d_sheet3, 1)
        copy_row(s_sheet, d_sheet4, 1)
    for item in filterList:
        count1=0
        count2=0
        count3=0
        
        for label in item.level:
            if label in label_1:
                count1 += 1
            elif label in label_2:
                count2 += 1
            elif label in label_3:
                count3 += 1
        # only level 1 rows
        if count1 != 0 and count2 == 0:
            for row in item.row_num:
                copy_row(s_sheet, d_sheet1, row)
            countL1 += 1
        # only level 2 rows
        elif count2 != 0 and count1 == 0:
            for row in item.row_num:
                copy_row(s_sheet, d_sheet2, row)
            countL2+=1
        # both level 1 and both level 2 items or split tags
        elif count3 !=0 or (count2 != 0 and count1 != 0):
            for row in item.row_num:
                copy_row(s_sheet, d_sheet3, row)
            countL3+=1
        # neigher of level 1 and level 2 items
        else :
            for row in item.row_num:
                copy_row(s_sheet, d_sheet4, row)
            countL4+=1
    set_label_counts_on_count_sheet(workbook[destination_sheet_5],
                                    destination_sheet_1,
                                    countL1,
                                    destination_sheet_2,
                                    countL2,
                                    destination_sheet_3,
                                    countL3,
                                    destination_sheet_4,
                                    countL4)
    # Saving after all the operations
    workbook.save(file_path)
    # Close the workbook
    workbook.close()

# name of variables
file_path = "file.xlsx"
source_sheet = "workingTab"
destination_sheet_1 = "Central"
destination_sheet_2 = "Unit"
destination_sheet_3 = "Split"
destination_sheet_4 = "Others"
destination_sheet_5 = "Counts"

# taking necessary searching data from requried sheet
get_required_values_from_one_tab(file_path, source_sheet)

copy_rows_to_dest_tab(file_path,
                      source_sheet, 
                      destination_sheet_1,
                      destination_sheet_2,
                      destination_sheet_3,
                      destination_sheet_4,
                      destination_sheet_5,
                      "Central",
                      "Unit",
                      "Split",
                      True)
print("Done")